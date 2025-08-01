import streamlit as st
import pandas as pd
from io import BytesIO, StringIO
import openpyxl
from openpyxl.utils import range_boundaries
from datetime import datetime

st.set_page_config(page_title="Cash Sales to IIF", layout="wide")
st.title("🧾 Convert Cash Sales Excel to QuickBooks IIF")

def unmerge_and_fill(workbook):
    for sheet in workbook.worksheets:
        for merged_range in list(sheet.merged_cells.ranges):
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            top_left_cell = sheet.cell(row=min_row, column=min_col)
            value = top_left_cell.value
            sheet.unmerge_cells(str(merged_range))
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.cell(row=row, column=col).value = value
    return workbook

def process_excel(file):
    try:
        in_memory_file = BytesIO(file.read())
        wb = openpyxl.load_workbook(in_memory_file)
        wb = unmerge_and_fill(wb)

        ws = wb.active
        data = ws.iter_rows(min_row=17, values_only=True)
        rows = list(data)

        # Define the columns we need
        col_till = 4    # column E (zero-indexed = 4)
        col_date = 9    # column J
        col_bill = 15   # column P
        col_amount = 25 # column Z

        # Extract rows until type stops being "Cash"
        extracted = []
        for row in rows:
            if not row or str(row[0]).strip().lower() != 'cash':
                break
            try:
                date_str = str(row[col_date]).strip()
                bill_date = pd.to_datetime(date_str)
            except Exception as e:
                continue
            extracted.append({
                'Till#': str(row[col_till]).strip(),
                'Bill Date': bill_date.strftime("%m/%d/%Y"),
                'Bill No.': str(row[col_bill]).strip(),
                'Amount': float(row[col_amount])
            })

        if not extracted:
            st.error("⚠️ No cash sales found.")
            return None, None

        df = pd.DataFrame(extracted)
        return df, generate_iif(df)

    except Exception as e:
        st.error(f"❌ Failed to process file: {e}")
        return None, None

def generate_iif(df):
    output = StringIO()
    output.write("!TRNS\tTRNSTYPE\tDATE\tACCNT\tNAME\tMEMO\tAMOUNT\tDOCNUM\n")
    output.write("!SPL\tTRNSTYPE\tDATE\tACCNT\tNAME\tMEMO\tAMOUNT\tQNTY\tINVITEM\n")
    output.write("!ENDTRNS\n")

    for _, row in df.iterrows():
        date = row['Bill Date']
        docnum = row['Bill No.']
        amount = row['Amount']
        till = row['Till#']

        memo = f"Till {till} - Bill {docnum}"

        output.write(f"TRNS\tPAYMENT\t{date}\tAccounts Receivable\tWalk In\t{memo}\t{amount:.2f}\t{docnum}\n")
        output.write(f"SPL\tPAYMENT\t{date}\tCash in Drawer\tWalk In\t{memo}\t{-amount:.2f}\t\t\n")
        output.write("ENDTRNS\n")

    return output.getvalue()

uploaded_file = st.file_uploader("📤 Upload Excel File", type=["xlsx"])
if uploaded_file:
    with st.spinner("Processing file..."):
        df, iif_content = process_excel(uploaded_file)

    if df is not None:
        st.success("✅ File processed successfully.")
        st.subheader("🔍 Preview of Extracted Data")
        st.dataframe(df.head(10))

        st.subheader("⬇️ Download QuickBooks IIF File")
        st.download_button(
            label="Download .IIF",
            data=iif_content,
            file_name="cash_sales.iif",
            mime="text/plain"
        )
